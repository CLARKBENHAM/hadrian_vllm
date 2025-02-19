# Used Adobe Acrobate to convert *fsi files into xlxs then munge tables together
# problem is unicode characters weren't copied around correctly
# %%
import os
import re
import openpyxl
import pandas as pd
import re

file_paths = [
    "data/fsi_labels/nist_ftc_06_asme1_rd_fsi.xlsx",
    "data/fsi_labels/nist_ftc_07_asme1_rd_fsi.xlsx",
    "data/fsi_labels/nist_ftc_08_asme1_rc1_fsi.xlsx",
    "data/fsi_labels/nist_ftc_09_asme1_rd_fsi.xlsx",
    "data/fsi_labels/nist_ftc_10_asme1_rb_fsi.xlsx",
    "data/fsi_labels/nist_ftc_11_asme1_rb_fsi.xlsx",
]

all_rows = []

for path in file_paths:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active

    base = os.path.basename(path)
    m = re.search(r"_(\d{2})_", base)
    assembly_id = m.group(1) if m else None

    header_row_index = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[:5] == (
            "Feature ID",
            "Feature Description",
            "Specification",
            "Element ID",
            "Comments",
        ):
            header_row_index = i
            break
    if not header_row_index:
        continue

    for row in ws.iter_rows(min_row=header_row_index + 1):
        relevant_cells = row[:5]
        any_colored = False
        for cell in relevant_cells:
            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.rgb:
                if fill.fgColor.rgb not in ("00000000", "FFFFFFFF"):
                    any_colored = True
                    break
        if any_colored and relevant_cells[0].value != "Feature ID":
            row_data = {
                "Feature ID": relevant_cells[0].value,
                "Feature Description": relevant_cells[1].value,
                "Specification": relevant_cells[2].value,
                "Element ID": relevant_cells[3].value,
                "Comments": relevant_cells[4].value,
                "Assembly ID": assembly_id,
            }
            all_rows.append(row_data)

df = pd.DataFrame(
    all_rows,
    columns=[
        "Feature ID",
        "Feature Description",
        "Specification",
        "Element ID",
        "Comments",
        "Assembly ID",
    ],
).map(lambda s: re.sub(r" +", " ", s) if isinstance(s, str) else s)
df.to_csv("data/fsi_labels/adobe_exported_all.csv", index=False)
df.to_clipboard(index=False)
